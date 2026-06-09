import csv
import gc
import io
import os
import tempfile
import unicodedata
from datetime import timedelta, datetime, date as date_type
from io import BytesIO

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.db.models import Count, Q, Prefetch
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl import Workbook
from .decorators import admin_or_coordinador_required, admin_required
from .forms import CreateDispatchGuideForm, UpdateGuideStateForm, ImportClientCSVForm, ImportDispatchExcelForm
from .utils import get_home_url_for_user, is_transportista, is_coordinador
from .models import Client, DispatchGuide, GuideStage, GuideStagePhoto, Seller

def home(request):
    if request.user.is_authenticated:
        return redirect(get_home_url_for_user(request.user))
    return redirect('login')


@admin_or_coordinador_required
def hub(request):
    resumen = {
        'total_guias': DispatchGuide.objects.count(),
        'en_ruta': DispatchGuide.objects.filter(estado='en_ruta').count(),
        'pendientes': DispatchGuide.objects.filter(estado__in=['emitida', 'asignada']).count(),
        'entregadas_hoy': DispatchGuide.objects.filter(
            estado='entregada',
            fecha_actualizacion__date=timezone.localdate()
        ).count(),
    }
    return render(request, 'guides/hub.html', {'resumen': resumen})


def user_login(request):
    if request.user.is_authenticated:
        return redirect('guide_list')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect(request.GET.get('next') or get_home_url_for_user(user))
    else:
        form = AuthenticationForm()

    return render(request, 'guides/login.html', {'form': form})


@login_required
def user_logout(request):
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('home')



@admin_or_coordinador_required
def guide_list(request):
    estado_filtro = request.GET.get('estado', '').strip()
    guides = DispatchGuide.objects.select_related('cliente', 'transportista', 'vendedor').order_by('-fecha_creacion')
    
    if estado_filtro:
        guides = guides.filter(estado=estado_filtro)
    
    context = {
        'guides': guides,
        'estado_filtro': estado_filtro,
        'estado_choices': DispatchGuide.STATUS_CHOICES,
    }
    return render(request, 'guides/guide_list.html', context)


@admin_or_coordinador_required
def export_route_planning(request):
    """Exporta un archivo Excel con la planificación de rutas."""
    guides = DispatchGuide.objects.select_related('cliente', 'transportista', 'vendedor').order_by('fecha_creacion')

    # Aplicar filtro por rango de fechas (fecha_envio)
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            guides = guides.filter(fecha_envio__gte=start_date)
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            guides = guides.filter(fecha_envio__lte=end_date)
    except ValueError:
        # Ignorar filtros mal formateados
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = 'Planificación Rutas'
    headers = [
        'Nota de Venta (NV)',
        'Creacion de la NV',
        'Fecha de Envio',
        'Rut Cliente',
        'Nombre cliente',
        'Direccion de despacho',
        'Persona que creo la NV',
        'Transportista asignado para esa nota de venta',
        'Fecha de despacho',
        'Numero de guia'
    ]
    ws.append(headers)

    for guide in guides:
        vendedor_nombre = guide.vendedor_nombre or (guide.vendedor.nombre if guide.vendedor else '')
        transportista_nombre = guide.transportista.get_full_name() if guide.transportista else ''
        ws.append([
            guide.nv or '',
            guide.nv_fecha_creacion or '',
            guide.fecha_envio or '',
            guide.cliente.rut,
            guide.cliente.nombre,
            guide.direccion_entrega,
            vendedor_nombre,
            transportista_nombre,
            guide.fecha_despacho or '',
            guide.numero_guia,
        ])

    for i, _ in enumerate(headers, start=1):
        column_letter = ws.cell(row=1, column=i).column_letter
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[column_letter]
        )
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=planificacion_rutas.xlsx'
    return response


@admin_or_coordinador_required
def dashboard(request):
    total_clients = Client.objects.count()
    total_guides = DispatchGuide.objects.count()
    total_stages = GuideStage.objects.count()
    status_counts = {status: 0 for status, _ in DispatchGuide.STATUS_CHOICES}
    for item in DispatchGuide.objects.values('estado').annotate(count=Count('id')):
        status_counts[item['estado']] = item['count']

    status_breakdown = []
    for status, label in DispatchGuide.STATUS_CHOICES:
        count = status_counts.get(status, 0)
        percent = (count / total_guides * 100) if total_guides else 0
        status_breakdown.append({
            'estado': status,
            'label': label,
            'count': count,
            'percent': round(percent, 1)
        })

    open_guides = DispatchGuide.objects.filter(~Q(estado__in=['entregada', 'rechazada', 'cerrada'])).count()
    closed_guides = DispatchGuide.objects.filter(estado__in=['entregada', 'rechazada', 'cerrada']).count()
    recent_guides = DispatchGuide.objects.select_related('cliente', 'transportista').order_by('-fecha_creacion')[:8]
    week_start = timezone.now() - timedelta(days=7)
    guides_last_7_days = DispatchGuide.objects.filter(fecha_creacion__gte=week_start).count()
    top_clients = Client.objects.annotate(guides_count=Count('guias')).order_by('-guides_count')[:5]
    top_carriers = User.objects.annotate(guides_count=Count('guias_asignadas')).filter(guides_count__gt=0).order_by('-guides_count')[:5]
    latest_stages = GuideStage.objects.select_related('guia').order_by('-timestamp')[:8]
    stage_photo_count = GuideStage.objects.filter(foto__isnull=False).count()

    context = {
        'total_clients': total_clients,
        'total_guides': total_guides,
        'total_stages': total_stages,
        'status_breakdown': status_breakdown,
        'open_guides': open_guides,
        'closed_guides': closed_guides,
        'recent_guides': recent_guides,
        'guides_last_7_days': guides_last_7_days,
        'top_clients': top_clients,
        'top_carriers': top_carriers,
        'latest_stages': latest_stages,
        'stage_photo_count': stage_photo_count,
    }
    return render(request, 'guides/dashboard.html', context)


@login_required
@require_POST
def search_client_by_rut(request):
    """API endpoint para buscar cliente por RUT (AJAX)."""
    rut = request.POST.get('rut', '').strip()
    
    if not rut:
        return JsonResponse({'error': 'RUT requerido'}, status=400)
    
    try:
        client = Client.objects.get(rut=rut)
        return JsonResponse({
            'success': True,
            'rut': client.rut,
            'nombre': client.nombre,
            'direccion_facturacion': client.direccion_facturacion,
            'direccion_entrega_preferida': client.direccion_entrega_preferida or ''
        })
    except Client.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': f'No existe cliente con RUT: {rut}'
        }, status=404)


@admin_or_coordinador_required
def create_guide(request):
    """Vista para crear una nueva guía de despacho (mobile-first)."""
    admin_session = request.user.is_staff or is_coordinador(request.user)

    if request.method == 'POST':
        form = CreateDispatchGuideForm(request.POST, request.FILES, admin_session=admin_session)

        rut = request.POST.get('rut', '').strip()
        direccion_entrega = request.POST.get('direccion_entrega', '').strip()
        usa_direccion_facturacion = request.POST.get('usa_direccion_facturacion') == 'on'
        map_link = request.POST.get('map_link', '').strip()
        
        if not rut:
            messages.error(request, 'Debe ingresar el RUT del cliente')
            return render(request, 'guides/create_guide.html', {'form': form, 'admin_session': admin_session})

        try:
            cliente = Client.objects.get(rut=rut)
        except Client.DoesNotExist:
            messages.error(request, f'No existe cliente con RUT: {rut}')
            return render(request, 'guides/create_guide.html', {'form': form, 'admin_session': admin_session})
        
        if form.is_valid():
            if not usa_direccion_facturacion and not direccion_entrega:
                messages.error(request, 'Debe ingresar una dirección de entrega si no usa la de facturación')
                return render(request, 'guides/create_guide.html', {'form': form, 'admin_session': admin_session})
            
            guide = form.save(commit=False)
            guide.cliente = cliente
            
            if usa_direccion_facturacion:
                guide.direccion_entrega = cliente.direccion_facturacion
            else:
                guide.direccion_entrega = direccion_entrega
            
            if map_link:
                guide.map_link = map_link
            
            if not admin_session:
                guide.vendedor_nombre = None
            guide.usa_direccion_facturacion = usa_direccion_facturacion
            guide.save()
            
            GuideStage.objects.create(
                guia=guide,
                estado='emitida',
                foto=form.cleaned_data.get('foto_emision'),
                observaciones='Guía emitida en el sistema'
            )
            
            messages.success(request, f'✓ Guía {guide.numero_guia} creada exitosamente')
            return redirect('guide_list')
    else:
        form = CreateDispatchGuideForm(admin_session=admin_session)
    
    context = {
        'form': form,
        'admin_session': admin_session,
    }
    return render(request, 'guides/create_guide.html', context)


@login_required
def transportista_guides(request):
    """Vista mobile-first para que el transportista vea sus guías asignadas."""
    if not is_transportista(request.user) and not request.user.is_staff:
        return redirect('guide_list')

    filtro = request.GET.get('filtro', 'activas')
    guides = DispatchGuide.objects.filter(
        transportista=request.user
    ).select_related('cliente').order_by('fecha_despacho', '-fecha_creacion')

    if filtro == 'activas':
        guides = guides.filter(estado__in=['emitida', 'asignada', 'en_ruta'])
    else:
        guides = guides.filter(estado__in=['entregada', 'rechazada', 'cerrada'])

    context = {
        'guides': guides,
        'filtro': filtro,
    }
    return render(request, 'guides/transportista_guides.html', context)


@login_required
def guide_detail(request, guide_id):
    """Vista para ver y actualizar el estado de una guía específica."""
    try:
        guide = DispatchGuide.objects.get(id=guide_id)
    except DispatchGuide.DoesNotExist:
        messages.error(request, 'Guía no encontrada')
        return redirect(get_home_url_for_user(request.user))

    # Transportista solo puede acceder a sus propias guías
    if is_transportista(request.user) and guide.transportista != request.user:
        messages.error(request, 'No tienes acceso a esta guía.')
        return redirect('transportista_guides')
    
    stages = guide.etapas.order_by('-timestamp')
    
    next_state_map = {
        'emitida': [('asignada', 'Asignar'), ('en_ruta', 'En Ruta'), ('rechazada', 'Rechazada')],
        'asignada': [('en_ruta', 'En Ruta'), ('entregada', 'Entregada'), ('rechazada', 'Rechazada')],
        'en_ruta': [('entregada', 'Entregada'), ('rechazada', 'Rechazada')],
        'entregada': [('cerrada', 'Cerrar')],
        'rechazada': [('cerrada', 'Cerrar')],
        'cerrada': [],
    }
    next_states = next_state_map.get(guide.estado, [])

    # Transportistas no pueden cerrar guías
    if is_transportista(request.user):
        next_states = [(s, l) for s, l in next_states if s != 'cerrada']
    
    if request.method == 'POST':
        form = UpdateGuideStateForm(request.POST, request.FILES)
        if form.is_valid():
            nuevo_estado = form.cleaned_data['nuevo_estado']
            evidencia_foto = form.cleaned_data.get('evidencia_foto')
            notas = form.cleaned_data.get('notas', '')

            fotos = request.FILES.getlist('evidencia_fotos')

            if nuevo_estado in ('entregada', 'rechazada') and not fotos:
                form.add_error('evidencia_foto', 'Debes subir al menos una foto al marcar como entregada o rechazada.')
            else:
                stage = GuideStage.objects.create(
                    guia=guide,
                    estado=nuevo_estado,
                    observaciones=notas
                )
                for i, foto in enumerate(fotos):
                    GuideStagePhoto.objects.create(etapa=stage, foto=foto, orden=i)

                guide.estado = nuevo_estado
                if nuevo_estado in ['entregada', 'rechazada'] and not guide.fecha_envio:
                    guide.fecha_envio = timezone.localdate()
                guide.save()

                messages.success(request, f'✓ Estado actualizado a "{guide.get_estado_display()}"')
                return redirect('guide_detail', guide_id=guide.id)
    else:
        form = UpdateGuideStateForm()
    
    template = (
        'guides/transportista_guide_detail.html'
        if is_transportista(request.user)
        else 'guides/guide_detail.html'
    )
    back_url = 'transportista_guides' if is_transportista(request.user) else 'guide_list'

    context = {
        'guide': guide,
        'stages': stages,
        'form': form,
        'next_states': next_states,
        'back_url': back_url,
    }
    return render(request, template, context)


@admin_or_coordinador_required
def transportista_report(request):
    from_date_str = request.GET.get('from_date', '')
    to_date_str = request.GET.get('to_date', '')

    guides_qs = DispatchGuide.objects.filter(
        transportista__isnull=False
    ).select_related('transportista', 'cliente').prefetch_related(
        Prefetch('etapas', queryset=GuideStage.objects.order_by('timestamp'))
    )

    try:
        if from_date_str:
            guides_qs = guides_qs.filter(
                fecha_creacion__date__gte=datetime.strptime(from_date_str, '%Y-%m-%d').date()
            )
        if to_date_str:
            guides_qs = guides_qs.filter(
                fecha_creacion__date__lte=datetime.strptime(to_date_str, '%Y-%m-%d').date()
            )
    except ValueError:
        pass

    transportistas_map = {}

    for guide in guides_qs:
        t = guide.transportista
        if t.id not in transportistas_map:
            nombre = t.get_full_name() or t.username
            transportistas_map[t.id] = {
                'transportista': t,
                'nombre': nombre,
                'por_despachar': 0,
                'en_ruta': 0,
                'entregadas': 0,
                'rechazadas': 0,
                'cerradas': 0,
                'total': 0,
                '_tiempos': [],
            }

        data = transportistas_map[t.id]
        data['total'] += 1

        if guide.estado in ('emitida', 'asignada'):
            data['por_despachar'] += 1
        elif guide.estado == 'en_ruta':
            data['en_ruta'] += 1
        elif guide.estado == 'entregada':
            data['entregadas'] += 1
        elif guide.estado == 'rechazada':
            data['rechazadas'] += 1
        elif guide.estado == 'cerrada':
            data['cerradas'] += 1

        etapas = list(guide.etapas.all())
        etapa_en_ruta = next((e for e in etapas if e.estado == 'en_ruta'), None)
        etapa_final = next((e for e in etapas if e.estado in ('entregada', 'rechazada')), None)
        if etapa_en_ruta and etapa_final:
            segundos = (etapa_final.timestamp - etapa_en_ruta.timestamp).total_seconds()
            if segundos > 0:
                data['_tiempos'].append(segundos)

    def fmt_time(seconds):
        h = int(seconds // 3600)
        m = int((seconds % 3600) // 60)
        return f"{h}h {m:02d}m"

    report_data = []
    for data in transportistas_map.values():
        tiempos = data.pop('_tiempos')
        data['completadas'] = data['entregadas'] + data['rechazadas']
        if tiempos:
            data['avg_time'] = fmt_time(sum(tiempos) / len(tiempos))
            data['min_time'] = fmt_time(min(tiempos))
            data['max_time'] = fmt_time(max(tiempos))
            data['guias_con_tiempo'] = len(tiempos)
        else:
            data['avg_time'] = data['min_time'] = data['max_time'] = None
            data['guias_con_tiempo'] = 0
        report_data.append(data)

    report_data.sort(key=lambda x: x['total'], reverse=True)

    totals = {
        'total':         sum(d['total'] for d in report_data),
        'por_despachar': sum(d['por_despachar'] for d in report_data),
        'en_ruta':       sum(d['en_ruta'] for d in report_data),
        'entregadas':    sum(d['entregadas'] for d in report_data),
        'rechazadas':    sum(d['rechazadas'] for d in report_data),
        'completadas':   sum(d['completadas'] for d in report_data),
    }

    return render(request, 'guides/transportista_report.html', {
        'report_data': report_data,
        'totals': totals,
        'from_date': from_date_str,
        'to_date': to_date_str,
    })


@admin_required
def import_clients(request):
    result = None

    if request.method == 'POST':
        form = ImportClientCSVForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo_csv']
            actualizar = form.cleaned_data['actualizar_existentes']

            try:
                texto = archivo.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    archivo.seek(0)
                    texto = archivo.read().decode('latin-1')
                except Exception:
                    messages.error(request, 'No se pudo leer el archivo. Verifica que sea UTF-8 o Latin-1.')
                    return render(request, 'guides/import_clients.html', {'form': form})

            # Detectar separador automáticamente
            muestra = texto[:2048]
            separador = ';' if muestra.count(';') >= muestra.count(',') else ','

            reader = csv.DictReader(io.StringIO(texto), delimiter=separador)

            # Normalizar nombres de columnas (strip + lower)
            campos_requeridos = {'rut', 'nombre', 'direccion_facturacion'}
            if not reader.fieldnames:
                messages.error(request, 'El archivo CSV está vacío o no tiene encabezados.')
                return render(request, 'guides/import_clients.html', {'form': form})

            columnas = {c.strip().lower(): c for c in reader.fieldnames}
            if not campos_requeridos.issubset(columnas.keys()):
                faltantes = campos_requeridos - set(columnas.keys())
                messages.error(request, f'Faltan columnas requeridas: {", ".join(faltantes)}')
                return render(request, 'guides/import_clients.html', {'form': form})

            creados = 0
            actualizados = 0
            errores = []

            for i, row in enumerate(reader, start=2):
                # Normalizar claves de la fila
                fila = {k.strip().lower(): (v.strip() if v else '') for k, v in row.items()}

                rut = fila.get('rut', '')
                nombre = fila.get('nombre', '')
                dir_facturacion = fila.get('direccion_facturacion', '')
                dir_entrega = fila.get('direccion_entrega_preferida', '') or None

                if not rut or not nombre or not dir_facturacion:
                    errores.append({
                        'fila': i,
                        'rut': rut or '—',
                        'motivo': 'Faltan campos obligatorios (rut, nombre o dirección de facturación)',
                    })
                    continue

                try:
                    cliente, created = Client.objects.get_or_create(
                        rut=rut,
                        defaults={
                            'nombre': nombre,
                            'direccion_facturacion': dir_facturacion,
                            'direccion_entrega_preferida': dir_entrega,
                        }
                    )
                    if created:
                        creados += 1
                    elif actualizar:
                        cliente.nombre = nombre
                        cliente.direccion_facturacion = dir_facturacion
                        cliente.direccion_entrega_preferida = dir_entrega
                        cliente.save()
                        actualizados += 1
                except Exception as e:
                    errores.append({'fila': i, 'rut': rut, 'motivo': str(e)})

            result = {
                'creados': creados,
                'actualizados': actualizados,
                'errores': errores,
                'total_procesados': creados + actualizados + len(errores),
            }
    else:
        form = ImportClientCSVForm()

    return render(request, 'guides/import_clients.html', {'form': form, 'result': result})


# ── Helpers para importación Excel ─────────────────────────────────────────

def _norm_col(name):
    """Normaliza nombre de columna: sin acentos, mayúsculas, sin espacios/puntos."""
    nfkd = unicodedata.normalize('NFKD', str(name))
    ascii_str = nfkd.encode('ASCII', 'ignore').decode('ASCII')
    return ascii_str.strip().upper().replace(' ', '_').replace('.', '').replace('°', '').replace('#', '')


def _find_col(headers, candidates):
    for c in candidates:
        if c in headers:
            return headers[c]
    return None


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    if isinstance(value, str):
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


@admin_or_coordinador_required
def bulk_assign_guides(request):
    transportistas = User.objects.filter(groups__name='Transportista').order_by('first_name', 'last_name')

    transportista_id = request.GET.get('transportista') or request.POST.get('transportista')
    transportista_sel = None
    guides = DispatchGuide.objects.none()

    if transportista_id:
        try:
            transportista_sel = User.objects.get(id=transportista_id, groups__name='Transportista')
            guides = DispatchGuide.objects.filter(
                estado__in=['emitida', 'asignada']
            ).select_related('cliente', 'transportista').order_by('fecha_despacho', '-fecha_creacion')
        except User.DoesNotExist:
            pass

    if request.method == 'POST' and transportista_sel:
        ids = request.POST.getlist('guia_ids')
        if ids:
            guias_a_asignar = DispatchGuide.objects.filter(id__in=ids)
            for guide in guias_a_asignar:
                era_emitida = guide.estado == 'emitida'
                guide.transportista = transportista_sel
                if era_emitida:
                    guide.estado = 'asignada'
                guide.save()
                if era_emitida:
                    GuideStage.objects.create(
                        guia=guide,
                        estado='asignada',
                        observaciones=f'Asignada a {transportista_sel.get_full_name() or transportista_sel.username}'
                    )
            messages.success(request, f'✓ {len(ids)} guía(s) asignadas a {transportista_sel.get_full_name() or transportista_sel.username}.')
            return redirect(f'{request.path}?transportista={transportista_sel.id}')

    context = {
        'transportistas': transportistas,
        'transportista_sel': transportista_sel,
        'guides': guides,
    }
    return render(request, 'guides/bulk_assign_guides.html', context)


# ── Vista de importación ────────────────────────────────────────────────────

def _process_sheet(ws, omitir_cr, actualizar):
    """Procesa la hoja en un único paso (compatible con read_only=True).
    Pre-carga lookups en memoria para evitar N+1 queries.
    Usa bulk_create/bulk_update para inserciones y actualizaciones masivas."""

    # ── Pre-cargar tablas de referencia (una query por tabla) ────────────
    clientes_map = {
        c.rut: c
        for c in Client.objects.only('id', 'rut', 'direccion_facturacion', 'direccion_entrega_preferida')
    }
    vendedores_map = {
        v.nombre.lower(): v
        for v in Seller.objects.filter(activo=True).only('id', 'nombre')
    }
    transportistas_list = list(
        User.objects.filter(groups__name='Transportista')
                    .only('id', 'first_name', 'last_name', 'username')
    )
    guias_existentes = set(DispatchGuide.objects.values_list('numero_guia', flat=True))

    # ── Única pasada sobre la hoja ───────────────────────────────────────
    FLUSH_EVERY = 100   # flush a BD cada N guías nuevas para limitar memoria pico
    col = {}
    header_found = False
    row_num = 0

    creadas = actualizadas = omitidas_cr = 0
    errores = []
    guides_nuevas  = []   # objetos DispatchGuide para bulk_create
    estados_nuevas = []   # estado inicial de cada guía nueva (mismo orden)
    updates_pendientes = {}  # {numero_guia: {campo: valor}}

    for raw_row in ws.iter_rows(values_only=True):
        row_num += 1

        if not any(v for v in raw_row if v is not None):
            continue

        # Primera fila con datos → encabezados
        if not header_found:
            headers = {}
            for idx, v in enumerate(raw_row):
                if v:
                    headers[_norm_col(str(v))] = idx
            col = {
                'guia':        _find_col(headers, ['GUIA', 'GUIA_DE_DESPACHO', 'N_GUIA']),
                'nv':          _find_col(headers, ['NV']),
                'creacion':    _find_col(headers, ['CREACION', 'FECHA_CREACION']),
                'fecha_envio': _find_col(headers, ['FENVIO', 'F_ENVIO', 'FECHA_ENVIO', 'ENVIO']),
                'rut_cliente': _find_col(headers, ['RUT_CLIENTE', 'RUTCLIENTE', 'RUT']),
                'referencia':  _find_col(headers, ['REFERENCIA']),
                'creado_por':  _find_col(headers, ['CREADO_POR', 'CREADOPOR', 'VENDEDOR']),
                'transporte':  _find_col(headers, ['TRANSPORTE']),
                'despacho':    _find_col(headers, ['DESPACHO', 'FECHA_DESPACHO']),
            }
            header_found = True
            continue

        def get(key, r=raw_row):
            idx = col.get(key)
            if idx is None or idx >= len(r):
                return None
            v = r[idx]
            return v.strip() if isinstance(v, str) else (v if v not in ('', None) else None)

        referencia = str(get('referencia') or '')
        if omitir_cr and 'cliente retira' in referencia.lower():
            omitidas_cr += 1
            continue

        numero_guia = get('guia')
        rut_raw     = get('rut_cliente')

        if not numero_guia:
            errores.append({'fila': row_num, 'guia': '—', 'motivo': 'Número de guía vacío'})
            continue
        numero_guia = str(numero_guia).strip()

        if not rut_raw:
            errores.append({'fila': row_num, 'guia': numero_guia, 'motivo': 'RUT cliente vacío'})
            continue

        cliente = clientes_map.get(str(rut_raw).strip())
        if not cliente:
            errores.append({'fila': row_num, 'guia': numero_guia, 'motivo': f'RUT {rut_raw} no existe en el sistema'})
            continue

        nv          = str(get('nv') or '').strip() or None
        nv_fecha    = _parse_date(get('creacion'))
        fecha_envio = _parse_date(get('fecha_envio'))
        fecha_desp  = _parse_date(get('despacho'))

        creado_por_raw = str(get('creado_por') or '').strip()
        vendedor_obj = vendedor_nombre = None
        if creado_por_raw:
            raw_lower = creado_por_raw.lower()
            for nombre_lower, v in vendedores_map.items():
                if raw_lower in nombre_lower or nombre_lower in raw_lower:
                    vendedor_obj = v
                    break
            if not vendedor_obj:
                vendedor_nombre = creado_por_raw

        transporte_raw = str(get('transporte') or '').strip().lower()
        transportista = None
        if transporte_raw:
            for t in transportistas_list:
                if (transporte_raw in (t.first_name or '').lower() or
                        transporte_raw in (t.last_name or '').lower() or
                        transporte_raw in t.username.lower()):
                    transportista = t
                    break

        # ── Guía existente ───────────────────────────────────────────────
        if numero_guia in guias_existentes:
            if actualizar:
                updates_pendientes[numero_guia] = {
                    'nv': nv,
                    'nv_fecha_creacion': nv_fecha,
                    'fecha_envio': fecha_envio,
                    'fecha_despacho': fecha_desp,
                    'notas': referencia or None,
                    'vendedor': vendedor_obj,
                    'vendedor_nombre': vendedor_nombre,
                    'transportista': transportista,
                }
            continue

        # ── Guía nueva ───────────────────────────────────────────────────
        estado_inicial = 'asignada' if transportista else 'emitida'
        guides_nuevas.append(DispatchGuide(
            numero_guia=numero_guia,
            nv=nv,
            nv_fecha_creacion=nv_fecha,
            cliente=cliente,
            direccion_entrega=cliente.direccion_entrega_preferida or cliente.direccion_facturacion,
            usa_direccion_facturacion=not bool(cliente.direccion_entrega_preferida),
            vendedor=vendedor_obj,
            vendedor_nombre=vendedor_nombre,
            transportista=transportista,
            fecha_envio=fecha_envio,
            fecha_despacho=fecha_desp,
            notas=referencia or None,
            estado=estado_inicial,
        ))
        estados_nuevas.append(estado_inicial)

        # Flush parcial: evita acumular cientos de objetos en RAM
        if len(guides_nuevas) >= FLUSH_EVERY:
            batch = DispatchGuide.objects.bulk_create(guides_nuevas, batch_size=FLUSH_EVERY)
            GuideStage.objects.bulk_create([
                GuideStage(guia=g, estado=s, observaciones='Importada desde Excel ERP')
                for g, s in zip(batch, estados_nuevas)
            ], batch_size=FLUSH_EVERY)
            creadas += len(batch)
            guides_nuevas.clear()
            estados_nuevas.clear()

    if not header_found:
        return None, 'La hoja seleccionada está vacía o no tiene encabezados.'

    # ── Bulk create del resto de guías nuevas (último batch) ─────────────
    if guides_nuevas:
        created = DispatchGuide.objects.bulk_create(guides_nuevas, batch_size=FLUSH_EVERY)
        GuideStage.objects.bulk_create([
            GuideStage(guia=g, estado=s, observaciones='Importada desde Excel ERP')
            for g, s in zip(created, estados_nuevas)
        ], batch_size=200)
        creadas = len(created)

    # ── Bulk update guías existentes ─────────────────────────────────────
    if updates_pendientes:
        guias_qs = list(DispatchGuide.objects.filter(numero_guia__in=updates_pendientes.keys()))
        changed_fields = set()
        for guide in guias_qs:
            data = updates_pendientes[guide.numero_guia]
            for field in ('nv', 'nv_fecha_creacion', 'fecha_envio', 'fecha_despacho',
                          'notas', 'vendedor', 'vendedor_nombre'):
                if data.get(field):
                    setattr(guide, field, data[field])
                    changed_fields.add(field)
            if data.get('transportista'):
                guide.transportista = data['transportista']
                changed_fields.add('transportista')
                if guide.estado == 'emitida':
                    guide.estado = 'asignada'
                    changed_fields.add('estado')
        if changed_fields:
            DispatchGuide.objects.bulk_update(guias_qs, fields=list(changed_fields), batch_size=200)
        actualizadas = len(guias_qs)

    return {
        'creadas': creadas,
        'actualizadas': actualizadas,
        'omitidas_cr': omitidas_cr,
        'errores': errores,
        'total': creadas + actualizadas + omitidas_cr + len(errores),
    }, None


@admin_or_coordinador_required
def import_guides_excel(request):
    form = ImportDispatchExcelForm()
    result = None

    if request.method == 'POST':
        action = request.POST.get('action', 'upload')

        # ── Paso 1: leer libros del archivo ─────────────────────────────
        if action == 'upload':
            form = ImportDispatchExcelForm(request.POST, request.FILES)
            if 'archivo_excel' not in request.FILES:
                messages.error(request, 'Selecciona un archivo Excel.')
                return render(request, 'guides/import_guides.html', {'form': form})

            archivo = request.FILES['archivo_excel']
            ext = os.path.splitext(archivo.name)[1].lower()

            if archivo.size > 10 * 1024 * 1024:
                messages.error(request, 'El archivo supera el límite de 10 MB. Divídelo en partes más pequeñas.')
                return render(request, 'guides/import_guides.html', {'form': form})

            try:
                with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                    for chunk in archivo.chunks():
                        tmp.write(chunk)
                    tmp_path = tmp.name

                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()

                request.session['excel_tmp_path'] = tmp_path
                request.session['excel_opts'] = {
                    'omitir_cr': request.POST.get('omitir_cliente_retira') == 'on',
                    'actualizar': request.POST.get('actualizar_existentes') == 'on',
                }

                return render(request, 'guides/import_guides.html', {
                    'form': form,
                    'step': 'select_sheet',
                    'sheets': sheets,
                })
            except Exception as e:
                if 'tmp_path' in locals() and os.path.exists(tmp_path):
                    os.unlink(tmp_path)
                messages.error(request, f'No se pudo leer el archivo: {e}')

        # ── Paso 2: importar la hoja seleccionada ───────────────────────
        elif action == 'import':
            tmp_path = request.session.pop('excel_tmp_path', None)
            opts = request.session.pop('excel_opts', {})
            hoja = request.POST.get('hoja', '')

            if not tmp_path or not os.path.exists(tmp_path):
                messages.error(request, 'La sesión expiró. Sube el archivo nuevamente.')
                return redirect('import_guides_excel')

            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb[hoja] if hoja in wb.sheetnames else wb.active
                result, error = _process_sheet(ws, opts.get('omitir_cr', True), opts.get('actualizar', False))
                wb.close()
                gc.collect()  # liberar memoria de los lookups y objetos post-importación
                if error:
                    messages.error(request, error)
            except Exception as e:
                messages.error(request, f'Error al procesar la hoja: {e}')
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

    return render(request, 'guides/import_guides.html', {'form': form, 'result': result})


@admin_or_coordinador_required
@require_POST
def update_guide_address(request, guide_id):
    guide = get_object_or_404(DispatchGuide, id=guide_id)
    direccion = request.POST.get('direccion_entrega', '').strip()
    map_link = request.POST.get('map_link', '').strip()

    if not direccion:
        return JsonResponse({'error': 'La dirección no puede estar vacía.'}, status=400)

    guide.direccion_entrega = direccion
    guide.map_link = map_link or None
    guide.save()

    return JsonResponse({
        'success': True,
        'direccion': direccion,
        'map_link': map_link,
    })
